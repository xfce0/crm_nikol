"""
Report Generators
PDF and Excel generation utilities

Note: This is a foundation. Install dependencies:
pip install reportlab openpyxl
"""

from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import io


# ============================================
# PDF GENERATOR
# ============================================

class PDFGenerator:
    """
    PDF report generator

    Requires: pip install reportlab
    """

    @staticmethod
    def generate_simple_report(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        output_path: str,
    ) -> str:
        """
        Generate a simple tabular PDF report

        Returns: Path to generated file
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            # Fallback: create a simple text file if reportlab not installed
            return PDFGenerator._create_text_fallback(title, data, headers, output_path)

        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.25 * inch))

        # Metadata
        meta_style = styles['Normal']
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        elements.append(Paragraph(f"Records: {len(data)}", meta_style))
        elements.append(Spacer(1, 0.5 * inch))

        # Table
        if data:
            # Prepare table data
            table_data = [headers]
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])

            # Create table
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("No data available", styles['Normal']))

        # Build PDF
        doc.build(elements)
        return output_path

    @staticmethod
    def _create_text_fallback(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        output_path: str,
    ) -> str:
        """Create a text fallback if reportlab is not installed"""
        with open(output_path.replace('.pdf', '.txt'), 'w') as f:
            f.write(f"{title}\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Records: {len(data)}\n\n")

            if headers and data:
                # Write headers
                f.write(" | ".join(headers) + "\n")
                f.write("-" * 80 + "\n")

                # Write data
                for row in data:
                    f.write(" | ".join(str(row.get(h, '')) for h in headers) + "\n")

        return output_path.replace('.pdf', '.txt')


# ============================================
# EXCEL GENERATOR
# ============================================

class ExcelGenerator:
    """
    Excel report generator

    Requires: pip install openpyxl
    """

    @staticmethod
    def generate_simple_report(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        output_path: str,
        sheet_name: str = "Report",
    ) -> str:
        """
        Generate a simple Excel report

        Returns: Path to generated file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback: create CSV if openpyxl not installed
            return ExcelGenerator._create_csv_fallback(title, data, headers, output_path)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:' + chr(65 + len(headers) - 1) + '1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Records: {len(data)}"

        # Headers (row 5)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row_idx, row in enumerate(data, start=6):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = chr(64 + col_idx)
            ws.column_dimensions[column_letter].width = max(len(header) + 5, 15)

        # Save
        wb.save(output_path)
        return output_path

    @staticmethod
    def _create_csv_fallback(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        output_path: str,
    ) -> str:
        """Create a CSV fallback if openpyxl is not installed"""
        import csv

        csv_path = output_path.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)

            # Write title as comment
            f.write(f"# {title}\n")
            f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Records: {len(data)}\n")

            writer.writeheader()
            writer.writerows(data)

        return csv_path


# ============================================
# REPORT TEMPLATES
# ============================================

class ReportTemplates:
    """Pre-defined report templates"""

    @staticmethod
    def projects_summary_template() -> Dict[str, Any]:
        """Project summary report template"""
        return {
            "title": "Projects Summary Report",
            "headers": ["ID", "Name", "Status", "Client", "Created", "Deadline", "Tasks Count"],
            "description": "Summary of all projects with key metrics",
        }

    @staticmethod
    def tasks_summary_template() -> Dict[str, Any]:
        """Task summary report template"""
        return {
            "title": "Tasks Summary Report",
            "headers": ["ID", "Title", "Project", "Assignee", "Status", "Priority", "Deadline"],
            "description": "Summary of all tasks",
        }

    @staticmethod
    def users_activity_template() -> Dict[str, Any]:
        """User activity report template"""
        return {
            "title": "Users Activity Report",
            "headers": ["User ID", "Username", "Email", "Role", "Tasks Assigned", "Tasks Completed", "Last Active"],
            "description": "User activity statistics",
        }

    @staticmethod
    def wb_products_template() -> Dict[str, Any]:
        """WildBerries products report template"""
        return {
            "title": "WildBerries Products Report",
            "headers": ["Article", "Name", "Price", "Stock", "Last Sync"],
            "description": "WildBerries products inventory",
        }


# ============================================
# EXPORT
# ============================================

__all__ = ["PDFGenerator", "ExcelGenerator", "ReportTemplates"]
